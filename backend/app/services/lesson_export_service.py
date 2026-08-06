import csv
import io
from datetime import datetime
from typing import Any, Dict, Iterable, List, Tuple, Union

import structlog

logger = structlog.get_logger()


class LessonExportService:

    CSV_HEADERS: List[str] = [
        "Lesson ID",
        "Class ID",
        "Class Name",
        "Grade Level",
        "Section",
        "Teacher ID",
        "Teacher Name",
        "Subject ID",
        "Subject Name",
        "Lesson Title",
        "Description",
        "Date",
        "Period",
        "Start Time",
        "End Time",
        "Status",
        "Visibility",
        "Seen Count",
        "Ack Count",
        "Coverage (Ack %)",
        "Homework Due Date",
        "Created At",
        "Updated At",
    ]

    @staticmethod
    def _coerce_to_string(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (list, dict)):
            try:
                import json
                return json.dumps(value, ensure_ascii=False)
            except Exception:
                return str(value)
        if isinstance(value, datetime):
            return value.isoformat()
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except Exception:
                pass
        return str(value)

    @staticmethod
    def _flatten_lesson(lesson: Any) -> Dict[str, Any]:
        lesson_id = getattr(lesson, "id", "")
        class_id = getattr(lesson, "class_id", "")
        class_obj = getattr(lesson, "class_", None)
        class_name = getattr(class_obj, "name", "") if class_obj else ""
        grade_level = getattr(class_obj, "grade_level", "") if class_obj else ""
        section = getattr(class_obj, "section", "") if class_obj else ""

        teacher_id = getattr(lesson, "teacher_id", "")
        teacher = getattr(lesson, "teacher", None)
        if teacher:
            teacher_name = (
                getattr(teacher, "full_name", None)
                or (
                    f"{getattr(teacher, 'first_name', '')} {getattr(teacher, 'last_name', '')}".strip()
                )
                or ""
            )
        else:
            teacher_name = ""

        subject_id = getattr(lesson, "subject_id", "")
        subject = getattr(lesson, "subject", None)
        subject_name = getattr(subject, "name", "") if subject else ""

        if not subject_name:
            materials = getattr(lesson, "materials", None) or []
            for m in materials:
                if isinstance(m, dict) and m.get("type") == "subject":
                    subject_name = m.get("subject_name") or m.get("value") or ""
                    break

        seen_count = int(getattr(lesson, "engagement_seen_count", 0) or 0)
        ack_count = int(getattr(lesson, "engagement_ack_count", 0) or 0)

        try:
            from app.models.student import Student
            class_enrollment = (
                Student.query.filter_by(class_id=class_id, status="active").count()
                if class_id
                else 0
            )
        except Exception:
            class_enrollment = 0
        coverage_pct = 0.0
        if class_enrollment:
            coverage_pct = round((ack_count / class_enrollment) * 100, 2)

        return {
            "Lesson ID": lesson_id,
            "Class ID": class_id,
            "Class Name": class_name,
            "Grade Level": grade_level,
            "Section": section,
            "Teacher ID": teacher_id,
            "Teacher Name": teacher_name,
            "Subject ID": subject_id,
            "Subject Name": subject_name,
            "Lesson Title": getattr(lesson, "title", ""),
            "Description": getattr(lesson, "description", ""),
            "Date": getattr(lesson, "date", ""),
            "Period": getattr(lesson, "period_number", ""),
            "Start Time": getattr(lesson, "start_time", ""),
            "End Time": getattr(lesson, "end_time", ""),
            "Status": getattr(lesson, "status", "") or "",
            "Visibility": getattr(lesson, "visibility", "") or "",
            "Seen Count": seen_count,
            "Ack Count": ack_count,
            "Coverage (Ack %)": coverage_pct,
            "Homework Due Date": getattr(lesson, "homework_due_date", ""),
            "Created At": getattr(lesson, "created_at", ""),
            "Updated At": getattr(lesson, "updated_at", ""),
        }

    @staticmethod
    def csv_export(filtered_lessons: Iterable[Any]) -> io.BytesIO:
        buffer = io.BytesIO()
        wrapper = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
        writer = csv.writer(wrapper)
        writer.writerow(LessonExportService.CSV_HEADERS)
        for lesson in filtered_lessons:
            flat = LessonExportService._flatten_lesson(lesson)
            row = [LessonExportService._coerce_to_string(flat.get(h, "")) for h in LessonExportService.CSV_HEADERS]
            writer.writerow(row)
        wrapper.flush()
        buffer.seek(0)
        return buffer

    @staticmethod
    def xlsx_export(filtered_lessons: Iterable[Any]) -> Union[io.BytesIO, Tuple[io.BytesIO, Dict[str, Any]]]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            xlsx_available = True
        except Exception:
            xlsx_available = False

        if not xlsx_available:
            csv_buf = LessonExportService.csv_export(filtered_lessons)
            warning_stub = {
                "warning": "openpyxl not importable; returning CSV binary inside a BytesIO wrapped as .xlsx download (stub fallback).",
                "actual_format": "text/csv",
                "generated_at": datetime.utcnow().isoformat(),
            }
            return csv_buf, warning_stub

        wb = Workbook()
        ws = wb.active
        ws.title = "Lessons"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0B1E35", end_color="0B1E35", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        ws.append(LessonExportService.CSV_HEADERS)
        for col_idx, _ in enumerate(LessonExportService.CSV_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for lesson in filtered_lessons:
            flat = LessonExportService._flatten_lesson(lesson)
            row_values = []
            for h in LessonExportService.CSV_HEADERS:
                raw = flat.get(h, "")
                if isinstance(raw, (int, float)):
                    row_values.append(raw)
                else:
                    row_values.append(LessonExportService._coerce_to_string(raw))
            ws.append(row_values)

        col_widths = [
            10, 9, 22, 12, 10, 10, 22, 10, 18, 30, 40, 12, 8, 11, 11,
            12, 14, 10, 10, 14, 16, 22, 22,
        ]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(LessonExportService.CSV_HEADERS))}{ws.max_row}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
